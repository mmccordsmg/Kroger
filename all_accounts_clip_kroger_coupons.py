from kroger_funcs import do_kroger_login, clip_kroger_coupons, check_kroger_coupon
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description="Clips kroger coupons in couponfile for accounts loaded from xlsx")
parser.add_argument("-i",	"--infile",			help="xlsx file with usernames in column D, passwords in column G",		required=True,	type=str)
parser.add_argument("-f", "--couponfile", help="File with coupon ids to clip", required=True, type=str)
args = parser.parse_args()

with open(args.couponfile, "r") as coupon_file:
	coupons_to_clip = coupon_file.read().splitlines()

wb = load_workbook(filename = args.infile)
for row in wb[wb.sheetnames[0]].values:
	username = row[3]
	password = row[6]
	if username is None or password is None:
		break
	if username == 'Email' or password == 'Password':
		continue
	#print("Username: %s Password: %s" % (username, password))
	loginresult = do_kroger_login(username, password)
	if not loginresult:
		print ("Error logging in to account %s" % username)
		continue
	print("Login successful! Clipping %d coupons on %s" % (len(coupons_to_clip),username))
	tosubmit = []
	for coupon_to_clip in coupons_to_clip:
		clipped, description = check_kroger_coupon(coupon_to_clip)
		if clipped:
			print("%s - %s already clipped on account %s" % (coupon_to_clip, description, username))
		else:
			tosubmit.append({"id": coupon_to_clip, "description": description})
	if tosubmit:
		clip_kroger_coupons(tosubmit)
	else:
		print("Nothing found to clip on account %s" % username)	